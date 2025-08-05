#!/usr/bin/env python3
"""
Test to verify everything works with the new HubSpot format including Excel headers
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from utils.parser import parse_ticket_data, HEADERS
from utils.excel import create_excel_file
from openpyxl import load_workbook

def test_new_format_excel():
    """Test new HubSpot format with Excel generation"""
    print("🧪 Testing New HubSpot Format with Excel Generation...")
    
    # New HubSpot format with blank line
    hubspot_data = """Cancellation of Steven Fricano's DNA Web App account

Preview
27333837927
Leon Morales
New (Support Pipeline)
Aug 1, 2025 10:14 PM GMT+5:30
Aug 1, 2025 10:48 PM GMT+5:30
Aug 1, 2025 10:14 PM GMT+5:30
Urgent
Isuru Promodh (isuru.promodh@dnabehavior.com)"""
    
    # Parse the data
    parsed_data, errors = parse_ticket_data(hubspot_data)
    
    print(f"✅ Parsing complete:")
    print(f"   - Tickets parsed: {len(parsed_data)}")
    print(f"   - Errors: {len(errors)}")
    
    if errors:
        print(f"   - Error details: {errors}")
        return False
    
    if not parsed_data:
        print("❌ No data parsed!")
        return False
    
    # Display parsed data
    print("\n📋 Parsed Data:")
    for key, value in parsed_data[0].items():
        print(f"   {key}: {value}")
    
    # Generate Excel file
    try:
        excel_buffer = create_excel_file(parsed_data)
        print(f"\n✅ Excel file generated successfully! ({len(excel_buffer.getvalue())} bytes)")
        
        # Verify Excel headers
        excel_buffer.seek(0)
        workbook = load_workbook(excel_buffer)
        worksheet = workbook.active
        
        excel_headers = []
        for col in range(1, len(HEADERS) + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            excel_headers.append(cell_value)
        
        if excel_headers == HEADERS:
            print("✅ Excel headers are in correct order!")
            
            # Check data row
            data_row = []
            for col in range(1, len(HEADERS) + 1):
                cell_value = worksheet.cell(row=2, column=col).value
                data_row.append(cell_value)
            
            print("\n📊 Excel Data Row 1:")
            for i, (header, value) in enumerate(zip(HEADERS, data_row)):
                print(f"   {header}: {value}")
            
            return True
        else:
            print("❌ Excel headers are in wrong order!")
            return False
            
    except Exception as e:
        print(f"❌ Excel generation failed: {e}")
        return False

if __name__ == "__main__":
    success = test_new_format_excel()
    if success:
        print("\n🎉 All tests passed! New format + Excel generation working perfectly!")
    else:
        print("\n❌ Some tests failed!")
        sys.exit(1)
