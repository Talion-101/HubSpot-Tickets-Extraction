"""
Excel export module for HubSpot ticket data
Creates Excel files in memory using openpyxl
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_excel_file(data):
    """
    Create an Excel file in memory from parsed ticket data
    
    Args:
        data (list): List of dictionaries with ticket data
        
    Returns:
        BytesIO: Excel file as bytes in memory
    """
    if not data:
        raise ValueError("No data provided for Excel export")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "HubSpot Tickets"
    
    # Get headers from first data row
    headers = list(data[0].keys())
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, ticket in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            value = ticket.get(header, '')
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        
        # Calculate max width needed
        max_length = len(header)
        for ticket in data:
            value_length = len(str(ticket.get(header, '')))
            if value_length > max_length:
                max_length = value_length
        
        # Set column width (with some padding)
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Apply borders and alignment to data cells
    from openpyxl.styles import Border, Side
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Apply styling to all cells
    for row in range(1, len(data) + 2):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if row > 1:  # Data rows (not header)
                cell.alignment = data_alignment
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def create_excel_with_pandas(data):
    """
    Alternative implementation using pandas (if preferred)
    
    Args:
        data (list): List of dictionaries with ticket data
        
    Returns:
        BytesIO: Excel file as bytes in memory
    """
    try:
        import pandas as pd
        
        if not data:
            raise ValueError("No data provided for Excel export")
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create BytesIO buffer
        excel_buffer = BytesIO()
        
        # Write to Excel with formatting
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='HubSpot Tickets', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['HubSpot Tickets']
            
            # Style the header row
            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        excel_buffer.seek(0)
        return excel_buffer
        
    except ImportError:
        # Fall back to openpyxl if pandas is not available
        return create_excel_file(data)
