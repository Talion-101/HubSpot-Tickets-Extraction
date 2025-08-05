#!/usr/bin/env python3
"""
Test to verify Excel header ordering is correct
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from utils.parser import parse_ticket_data, HEADERS
from utils.excel import create_excel_file
from openpyxl import load_workbook
from io import BytesIO

def test_excel_header_order():
    """Test that Excel headers are in the correct order"""
    print("🧪 Testing Excel Header Order...")
    
    # Test data
    test_data = """Login Issue | TK-001 | john.doe@example.com | Open | 2025-08-01 | 2025-08-04 | 2025-08-03 | High | Alice
Payment Failure | TK-002 | jane.smith@example.com | In Progress | 2025-07-29 | 2025-08-02 | 2025-08-01 | Medium | Bob"""
    
    # Parse the data
    parsed_data, errors = parse_ticket_data(test_data)
    
    if errors:
        print(f"❌ Parsing errors: {errors}")
        return False
    
    # Generate Excel file
    excel_buffer = create_excel_file(parsed_data)
    
    # Load the workbook to check headers
    excel_buffer.seek(0)  # Reset buffer position
    workbook = load_workbook(excel_buffer)
    worksheet = workbook.active
    
    # Get headers from Excel file
    excel_headers = []
    for col in range(1, len(HEADERS) + 1):
        cell_value = worksheet.cell(row=1, column=col).value
        excel_headers.append(cell_value)
    
    print(f"Expected headers: {HEADERS}")
    print(f"Excel headers:    {excel_headers}")
    
    # Check if headers match expected order
    if excel_headers == HEADERS:
        print("✅ Excel headers are in correct order!")
        return True
    else:
        print("❌ Excel headers are NOT in correct order!")
        return False

if __name__ == "__main__":
    success = test_excel_header_order()
    if success:
        print("\n🎉 Excel header order test passed!")
    else:
        print("\n❌ Excel header order test failed!")
        sys.exit(1)
